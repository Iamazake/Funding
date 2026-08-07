from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.excel.contract import APPROVED_COLUMNS
from app.services.excel.mapping import WorkbookMapper
from app.services.excel.reader import OperationalWorkbookReader
from app.services.excel.source import LocalFileSource
from tests.excel_helpers import VALID_CPF, create_workbook, default_rows


def map_workbook(path: Path):
    with LocalFileSource(path).stage() as staged:
        with OperationalWorkbookReader(staged) as reader:
            reader.validate_layout()
            return WorkbookMapper().map(reader)


def test_orphans_are_imported_and_reported(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        rows={
            "DFEN_CONTRATO": {"COD_CLIENTE": "SEM-CLIENTE", "NUM_CPF": "11111111111"},
            "ECON_EMPRESTIMOS": {"COD_CONTRATO": "SEM-CONTRATO"},
            "ECON_AMORTIZACOES": {"COD_CONTRATO": "SEM-PAIS"},
        },
    )
    imported = map_workbook(original)
    issue_types = {issue.inconsistency_type for issue in imported.inconsistencies}

    assert "orphan_contract" in issue_types
    assert "orphan_loan" in issue_types
    assert "orphan_amortization_contract" in issue_types
    assert "orphan_amortization_loan" in issue_types
    assert len(imported.rows["DFEN_CONTRATO"]) == 1
    assert len(imported.rows["ECON_EMPRESTIMOS"]) == 1
    assert len(imported.rows["ECON_AMORTIZACOES"]) == 1


def test_secondary_invalid_fields_become_warning_and_preserve_raw_value(tmp_path: Path) -> None:
    bad_client = {
        "COD_CLIENTE": "CLI-2",
        "CPF_CLIENTE": "111.111.111-11",
        "NOME_CLIENTE": "Linha Inválida",
        "DT_NASC": "31/02/2026",
    }
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        extra_rows={"BCLI_CADASTRO": [bad_client]},
    )
    imported = map_workbook(original)
    rows = imported.rows["BCLI_CADASTRO"]

    assert len(rows) == 2
    assert rows[0].validation_status == "valid"
    assert rows[1].validation_status == "warning"
    assert rows[1].data["cpf_normalized"] is None
    assert rows[1].data["dt_nasc"] is None
    assert rows[1].raw_data["CPF_CLIENTE"] == "111.111.111-11"
    assert imported.counters()["sheets"]["BCLI_CADASTRO"]["warning"] == 1
    assert imported.counters()["sheets"]["BCLI_CADASTRO"]["invalid"] == 0


def test_multiple_payment_movements_are_preserved_as_information(tmp_path: Path) -> None:
    duplicate = default_rows()["ECON_AMORTIZACOES"]
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        extra_rows={"ECON_AMORTIZACOES": [duplicate]},
    )
    imported = map_workbook(original)

    assert len(imported.rows["ECON_AMORTIZACOES"]) == 2
    assert any(
        issue.inconsistency_type == "multiple_payment_movements"
        and issue.severity == "info"
        for issue in imported.inconsistencies
    )
    assert all(row.validation_status == "valid" for row in imported.rows["ECON_AMORTIZACOES"])


def test_baixa_total_is_raw_marker_and_never_money(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        rows={"ECON_AMORTIZACOES": {"BAIXA _TOTAL": "S"}},
    )
    imported = map_workbook(original)
    row = imported.rows["ECON_AMORTIZACOES"][0]

    assert row.data["baixa_total_original"] == "S"
    assert "baixa_total" not in row.data
    assert not any(
        issue.field_name == "BAIXA _TOTAL" and issue.inconsistency_type == "ambiguous_money"
        for issue in imported.inconsistencies
    )
    assert row.validation_status == "valid"


def test_all_parsed_numbers_are_decimal_or_integer_never_float(tmp_path: Path) -> None:
    original = create_workbook(tmp_path / "Cadastro de Clientes.xlsm")
    imported = map_workbook(original)

    def assert_no_float(value: Any) -> None:
        assert not isinstance(value, float)
        if isinstance(value, dict):
            for nested in value.values():
                assert_no_float(nested)
        elif isinstance(value, list):
            for nested in value:
                assert_no_float(nested)

    for rows in imported.rows.values():
        for row in rows:
            assert_no_float(row.data)
            assert_no_float(row.raw_data)
    loan = imported.rows["ECON_EMPRESTIMOS"][0]
    assert isinstance(loan.data["vl_principal"], Decimal)
    assert isinstance(loan.data["taxa_juros"], Decimal)


def test_valid_cpf_is_normalized_in_every_sheet(tmp_path: Path) -> None:
    original = create_workbook(tmp_path / "Cadastro de Clientes.xlsm")
    imported = map_workbook(original)
    assert imported.rows["BCLI_CADASTRO"][0].data["cpf_normalized"] == VALID_CPF
    assert imported.rows["DFEN_CONTRATO"][0].data["cpf_normalized"] == VALID_CPF
    assert imported.rows["ECON_EMPRESTIMOS"][0].data["cpf_normalized"] == VALID_CPF
    assert imported.rows["ECON_AMORTIZACOES"][0].data["cpf_normalized"] == VALID_CPF


def test_repeated_client_rows_are_preserved_without_deduplication(tmp_path: Path) -> None:
    repeated = default_rows()["BCLI_CADASTRO"]
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        extra_rows={"BCLI_CADASTRO": [repeated]},
    )
    imported = map_workbook(original)
    rows = imported.rows["BCLI_CADASTRO"]
    assert len(rows) == 2
    assert rows[0].source_key == rows[1].source_key


def test_raw_data_contains_only_allowlisted_columns(tmp_path: Path) -> None:
    original = create_workbook(tmp_path / "Cadastro de Clientes.xlsm")
    workbook = load_workbook(original)
    worksheet = workbook["BCLI_CADASTRO"]
    extra_column = worksheet.max_column + 1
    worksheet.cell(row=1, column=extra_column, value="CAMPO_FORA_DA_ALLOWLIST")
    worksheet.cell(row=2, column=extra_column, value="não pode persistir")
    workbook.save(original)
    workbook.close()

    imported = map_workbook(original)
    raw_data = imported.rows["BCLI_CADASTRO"][0].raw_data
    assert set(raw_data) == set(APPROVED_COLUMNS["BCLI_CADASTRO"])
    assert "CAMPO_FORA_DA_ALLOWLIST" not in raw_data


def test_duplicate_contract_codes_are_registered_without_deduplication(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        extra_rows={
            "DFEN_CONTRATO": [default_rows()["DFEN_CONTRATO"]],
            "ECON_EMPRESTIMOS": [default_rows()["ECON_EMPRESTIMOS"]],
        },
    )
    imported = map_workbook(original)
    issue_types = [issue.inconsistency_type for issue in imported.inconsistencies]

    assert len(imported.rows["DFEN_CONTRATO"]) == 2
    assert len(imported.rows["ECON_EMPRESTIMOS"]) == 2
    assert issue_types.count("duplicate_dfen_contract_code") == 1
    assert issue_types.count("duplicate_loan_contract_code") == 1


def test_orphans_are_divergent_not_invalid(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        rows={
            "ECON_EMPRESTIMOS": {"COD_CONTRATO": "ORPHAN-LOAN"},
            "ECON_AMORTIZACOES": {"COD_CONTRATO": "ORPHAN-INSTALLMENT"},
        },
    )
    imported = map_workbook(original)

    assert imported.rows["ECON_EMPRESTIMOS"][0].validation_status == "divergent"
    assert imported.rows["ECON_AMORTIZACOES"][0].validation_status == "divergent"


def test_secondary_money_error_is_warning(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        rows={"ECON_AMORTIZACOES": {"DESCONTO_CONC": "valor ambíguo"}},
    )
    imported = map_workbook(original)
    row = imported.rows["ECON_AMORTIZACOES"][0]

    assert row.validation_status == "warning"
    assert row.data["desconto_conc"] is None


def test_missing_essential_identifier_is_invalid(tmp_path: Path) -> None:
    original = create_workbook(
        tmp_path / "Cadastro de Clientes.xlsm",
        rows={"ECON_AMORTIZACOES": {"COD_CONTRATO": None}},
    )
    imported = map_workbook(original)
    row = imported.rows["ECON_AMORTIZACOES"][0]

    assert row.validation_status == "invalid"
    assert any(
        issue.inconsistency_type == "missing_required_identifier"
        and issue.severity == "invalid"
        for issue in imported.inconsistencies
    )
