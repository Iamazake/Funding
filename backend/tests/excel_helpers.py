from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.services.excel.contract import APPROVED_COLUMNS, AUTHORIZED_SHEETS

VALID_CPF = "52998224725"


def default_rows() -> dict[str, dict[str, Any]]:
    return {
        "BCLI_CADASTRO": {
            "COD_CLIENTE": "CLI-1",
            "CPF_CLIENTE": VALID_CPF,
            "NOME_CLIENTE": "Cliente Sintético",
            "DT_NASC": "01/02/1990",
        },
        "DFEN_CONTRATO": {
            "COD_CLIENTE": "CLI-1",
            "COD_CONTRATO": "CTR-1",
            "NUM_CPF": VALID_CPF,
            "DT_OPERACAO": "01/01/2026",
            "VCTO_PRIM_PARC": "01/02/2026",
            "PRAZO": 12,
            "PRINCIPAL": "1.234,56",
            "IOF": "12,34",
            "VL_FINANCIADO": "1246.90",
            "PMT": "103,91",
            "VL_LIBERADO": 1200,
            "DATA_LIBERACAO": "02/01/2026",
        },
        "ECON_EMPRESTIMOS": {
            "COD_CONTRATO": "CTR-1",
            "COD_CLIENTE": "CLI-1",
            "NUM_CPF": VALID_CPF,
            "DT_OPERACAO": "01/01/2026",
            "VENCIMENTO1": "01/02/2026",
            "VL_PRINCIPAL": 1234.56,
            "PRAZO_PGTO": 12,
            "IOF": 12.34,
            "VL_FINACIADO": 1246.90,
            "PMT": 103.91,
            "VL_LIBERADO": 1200,
            "TAXA_JUROS": 0.02,
            "TAXA_TIR": 0.021,
            "TAXA_CET_AM": 0.023,
            "STATUS": "ATIVO",
        },
        "ECON_AMORTIZACOES": {
            "COD_CLIENTE": "CLI-1",
            "NUM_CPF": VALID_CPF,
            "COD_CONTRATO": "CTR-1",
            "COD_PARCELA": 1,
            "VENCIMENTO": "01/02/2026",
            "VAL_AMTZ_JUR": 20.50,
            "VAL_AMTZ_PRINC": 83.41,
            "VAL_PARCELA": 103.91,
            "BAIXA _TOTAL": 0,
            "DT_BAIXATOTAL": None,
            "VAL_PGTO": 0,
            "DESCONTO_CONC": 0,
            "STATUS_PARC": "ABERTA",
            "SITUACAO": "A VENCER",
            "BOL_ANTECIP": "NÃO",
            "PRODUTO_FINANCEIRO": "SINTÉTICO",
        },
    }


def create_workbook(
    path: Path,
    *,
    rows: Mapping[str, Mapping[str, Any]] | None = None,
    missing_sheet: str | None = None,
    missing_column: tuple[str, str] | None = None,
    extra_rows: Mapping[str, list[Mapping[str, Any]]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    source_rows = default_rows()
    if rows:
        for sheet, overrides in rows.items():
            source_rows[sheet].update(overrides)

    for sheet_name in AUTHORIZED_SHEETS:
        if sheet_name == missing_sheet:
            continue
        worksheet = workbook.create_sheet(sheet_name)
        headers = list(APPROVED_COLUMNS[sheet_name])
        if missing_column and missing_column[0] == sheet_name:
            headers.remove(missing_column[1])
        worksheet.append(headers)
        worksheet.append([source_rows[sheet_name].get(header) for header in headers])
        for row in (extra_rows or {}).get(sheet_name, []):
            worksheet.append([row.get(header) for header in headers])

    sensitive = workbook.create_sheet("CAD_USUARIOS")
    sensitive.append(["USUARIO", "SENHA"])
    sensitive.append(["nunca-ler", "segredo-sintetico"])
    workbook.create_sheet("Planilha1")
    workbook.create_sheet("BCLI_DADOS_BANCARIOS")
    workbook.create_sheet("NAO_AUTORIZADA")
    workbook.save(path)
    workbook.close()
    return path
