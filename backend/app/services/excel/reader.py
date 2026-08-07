from __future__ import annotations

from collections.abc import Iterator, Mapping
from contextlib import AbstractContextManager
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.services.excel.contract import (
    APPROVED_COLUMNS,
    AUTHORIZED_SHEETS,
    OPTIONAL_COLUMNS,
    SENSITIVE_SHEETS,
)
from app.services.excel.errors import (
    MissingRequiredColumnError,
    MissingRequiredSheetError,
    SensitiveSheetAccessError,
    UnauthorizedSheetError,
    WorkbookAccessError,
)
from app.services.excel.source import StagedFile


@dataclass(frozen=True, slots=True)
class SheetRow:
    number: int
    values: Mapping[str, Any]


class OperationalWorkbookReader(AbstractContextManager["OperationalWorkbookReader"]):
    def __init__(self, staged_file: StagedFile) -> None:
        if not staged_file.is_reader_approved():
            raise WorkbookAccessError("O reader aceita somente uma cópia temporária aprovada.")
        self._staged_file = staged_file
        self._workbook: Workbook | None = None
        self._column_indexes: dict[str, dict[str, int]] = {}

    @property
    def epoch(self):
        return self._require_workbook().epoch

    def __enter__(self) -> OperationalWorkbookReader:
        self._workbook = load_workbook(
            filename=self._staged_file.copy_path,
            read_only=True,
            data_only=True,
            keep_vba=False,
            keep_links=False,
        )
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def validate_layout(self) -> None:
        workbook = self._require_workbook()
        available = set(workbook.sheetnames)
        for sheet_name in AUTHORIZED_SHEETS:
            if sheet_name not in available:
                raise MissingRequiredSheetError(f"A aba obrigatória {sheet_name} está ausente.")

        # Worksheet objects are requested only for the four positive-list sheets.
        for sheet_name in AUTHORIZED_SHEETS:
            worksheet = workbook[sheet_name]
            header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            indexes: dict[str, int] = {}
            for index, value in enumerate(header_values):
                normalized = str(value).strip() if value is not None else ""
                if normalized and normalized not in indexes:
                    indexes[normalized] = index
            missing = [column for column in APPROVED_COLUMNS[sheet_name] if column not in indexes]
            if missing:
                names = ", ".join(missing)
                raise MissingRequiredColumnError(
                    f"A aba {sheet_name} não possui coluna(s) obrigatória(s): {names}."
                )
            self._column_indexes[sheet_name] = indexes

    def iter_sheet(self, sheet_name: str) -> Iterator[SheetRow]:
        self._assert_authorized(sheet_name)
        if sheet_name not in self._column_indexes:
            raise WorkbookAccessError("O layout deve ser validado antes da leitura.")

        workbook = self._require_workbook()
        worksheet = workbook[sheet_name]
        requested = APPROVED_COLUMNS[sheet_name] + OPTIONAL_COLUMNS.get(sheet_name, ())
        indexes = self._column_indexes[sheet_name]
        selected = tuple((column, indexes[column]) for column in requested if column in indexes)
        max_index = max(index for _, index in selected)
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=max_index + 1, values_only=True), start=2
        ):
            values = {column: row[index] for column, index in selected}
            if any(value is not None and value != "" for value in values.values()):
                yield SheetRow(row_number, values)

    def _assert_authorized(self, sheet_name: str) -> None:
        if sheet_name in SENSITIVE_SHEETS:
            raise SensitiveSheetAccessError(f"A aba sensível {sheet_name} está bloqueada.")
        if sheet_name not in AUTHORIZED_SHEETS:
            raise UnauthorizedSheetError(f"A aba {sheet_name} não está autorizada.")

    def _require_workbook(self) -> Workbook:
        if self._workbook is None:
            raise WorkbookAccessError("O workbook temporário não está aberto.")
        return self._workbook
